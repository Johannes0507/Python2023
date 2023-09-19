# -*- coding: utf-8 -*-
"""
Created on Sat May 13 22:43:15 2023

@author: 李旻恩
"""

import openpyxl

file = 'school.xlsx'

# data_only=True 只抓資料
wb = openpyxl.load_workbook(file,data_only=True)

wb = wb['students']

for i in range(1,ws.max_column + 1):
    print(ws.cell(column=i,row=5).value,end=' ')
    
print()

for i in range(2,ws.max_row + 1):
    for y in range(1,ws.max_column + 1):
        print(ws.cell(i,y).value,end='')
    print()
    
print('第二種寫法')

for row in ws.rows:
    for cell in row:
        print(cell.value,end=' ')
    print()
    
