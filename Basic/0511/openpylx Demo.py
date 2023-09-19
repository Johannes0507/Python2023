# -*- coding: utf-8 -*-
"""
Created on Sat May 13 21:47:52 2023

@author: 李旻恩
"""
import openpyxl

file = 'school.xlsx'

wb = openpyxl.load_workbook(file)

# 抓取預設工作表名稱
ws = wb.active

print('目前工作表：',ws.title)

ws = wb['students']

print('所選工作預設表：',ws.title)

print('A2儲存格值：',ws['A2'].value)
print('B2儲存格值：',ws['B2'].value)
print('C2儲存格值：',ws['C2'].value)

# 找位置
print(ws['A2'].column,ws['A2'].row)
print(ws['C1'].column,ws['C1'].row)

# 有資料的總攔
print('工作表攔數：',ws.max_column)
# 有資料的資料總列
print('工作表列數：',ws.max_row)