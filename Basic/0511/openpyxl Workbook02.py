# -*- coding: utf-8 -*-
"""
Created on Sun May 14 22:53:35 2023

@author: 李旻恩
"""

import openpyxl

wb = openpyxl.Workbook()

ws = wb.active

ws.title = 'member'

# 第一種
ws['A1'] = '姓名'
ws['A2'] = 'Bill'
ws['B1'] = '學校'
ws['B2'] = '聯成'

# 第二種
rowdata = ['David','中科大']
ws.append(rowdata)

# 第三種
data = [['John','中興']['Mary','靜宜']['Sue','東海']]
for item in data:
    ws.append(item)

wb.save('lcc2.xlsx')